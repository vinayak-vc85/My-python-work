""" from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import colors, PatternFill  """
import pandas as pd

class style_excel:
    def __init__(self, filename='data_difference.xlsx'):
        self.filename = filename
    
    def color_negative_red(self,val):
    
        if '-->' in str(val):
        
            color = 'red' 
            
        else:
            color = 'black'
        return 'color: %s' % color
    def colorbg(self,val):
        if '-->' in str(val):
        
            color = 'yellow' 
            
        else:
            color = 'white'
        return 'background-color: %s' % color

    def perform_style(self):
        df = pd.read_excel(self.filename,  na_values='None' ,sheet_name='All Difference')
        s = df.style.applymap(self.color_negative_red).applymap(self.colorbg)
        s.to_excel(r'difference_style.xlsx', index = False)
        return "Styles applied for All Difference"

'''''
workbook = load_workbook(filename="data_difference.xlsx")
sheet = workbook["All Difference"]
bg = PatternFill(bgColor="a832a6")
rule = CellIsRule(operator='containsText', formula=[' -->'], fill=bg)
sheet.conditional_formatting.add( "A1:Z104",rule)
workbook.save('wb2b.xlsx') '''

